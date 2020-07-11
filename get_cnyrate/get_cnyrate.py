
import requests
from bs4 import BeautifulSoup

rates = []#empty list

#exchangers
html = requests.get("https://www.exchangers.co.jp/rate.php")
soup = BeautifulSoup(html.content, "html.parser")

det = 0
for detail in soup.find_all(class_="common-currency"):
    if det == 2:#人民元は3番目
        ret = detail.find_all(class_="bdt")
        #shop_a = float(ret[2].text)
        rates.append(float(ret[2].text))
    det += 1
    #print(ret[1].text)
    #print(ret[2].text)


#D-ranger
html = requests.get("https://d-ranger.jp/shop/shinjuku/")
soup = BeautifulSoup(html.content, "html.parser")

det = 0
for detail in soup.find_all("tr"):
    if detail.find(scope="row"):
        if det == 2:#人民元は3番目
            #shop_b = float(detail.find(class_="cell-buy").text[:-1])
            rates.append(float(detail.find(class_="cell-buy").text[:-1]))
        det += 1
        #print(detail.find(class_="shoprate-name").text)
        #print(detail.find(class_="cell-buy").text)


#interbank
html = requests.get("https://www.interbank.co.jp/")
soup = BeautifulSoup(html.content, "html.parser")

det = 0
for detail in soup.find_all(class_="subBox"):
    if det == 2:#人民元は3番目
        ret = detail.find(class_="rBox")
        #shop_c = float(ret.find("dt").text)
        rates.append(float(ret.find("dt").text))
    det += 1
    #print(detail.find("h3").text)
    #ret = detail.find(class_="rBox")
    #print(ret.find("dt").text)


#print("Compare CNY Rates\nExchangers:%.2f\nD-ranger:%.2f\nInterbank:%.2f"%(shop_a,shop_b,shop_c))
#print("Compare CNY Rates\nExchangers:%.2f\nD-ranger:%.2f\nInterbank:%.2f"%(rates[0],rates[1],rates[2]))

import openpyxl
import datetime

# ワークブックを新規作成する
#book = openpyxl.Workbook()
book = openpyxl.load_workbook("demo.xlsx")

# シートを取得し名前を変更する
sheet = book.active
#sheet.title = 'First sheet'
buf = int(sheet.max_row)

sheet['A'+str(buf+1)] = datetime.datetime.now()
sheet['B'+str(buf+1)] = rates[0]#a2,a3と継ぎ足ししたい
sheet['C'+str(buf+1)] = rates[1]
sheet['D'+str(buf+1)] = rates[2]

# ワークブックに名前をつけて保存する
book.save('demo.xlsx')


#reference
#https://codezine.jp/article/detail/12230
#https://tonari-it.com/python-html-get-text-attr/#toc8